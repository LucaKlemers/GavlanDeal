from django.shortcuts import render, HttpResponse
from integration_utils.bitrix24.bitrix_user_auth.main_auth import main_auth
from integration_utils.bitrix24.functions.call_list_method import call_list_method
from integration_utils.bitrix24.functions.batch_api_call import _batch_api_call
from django.conf import settings
from operator import itemgetter
from django.core.signing import Signer
from django.utils.crypto import get_random_string
import qrcode
from .models import Product #, ProductID
import requests
from PIL import Image
import io
import base64
from functools import lru_cache
from datetime import datetime, timedelta, timezone
from os.path import splitext
import csv
import openpyxl

@main_auth(on_start=True, set_cookie=True)
def index(request):
    app_settings = settings.APP_SETTINGS
    user = request.bitrix_user
    user_name = user.first_name
    context = {"app_settings": app_settings, "user_name": user_name,}
    return render(request, 'GavlanDealApp/index.html', context)


@main_auth(on_cookies=True)
def last_deals(request):
    token = request.bitrix_user_token
    # full_deals = sorted(call_list_method(token, 'crm.deal.list', ['select => ["ID", "TITLE", "OPPORTUNITY", "TAX_VALUE", "BEGINDATE", "CLOSEDATE", "UF_CRM_1752242366887"]']), key=itemgetter('LAST_ACTIVITY_TIME'), reverse=True)[:10]
    fields = ["ID", "Наименование", "Валюта", "Сумма", "Налог",
             # "Клиент",
              "Начало", "Конец",  "Скидка(%)"]
    deal_types = ["ID", "TITLE", "CURRENCY_ID", "OPPORTUNITY", "TAX_VALUE", "BEGINDATE", "CLOSEDATE", "UF_CRM_1752242366887"]
    full_deals = sorted(token.call_api_method("crm.deal.list")['result'], key=itemgetter('LAST_ACTIVITY_TIME'), reverse=True)[:10]
    deals = []
    for deal in full_deals:
        id= {'ID': deal["ID"]}
        real_deal = token.call_api_method('crm.deal.get', params=id)['result']
        values = []
        for value in real_deal:
            if value in deal_types:
                values.append(real_deal[value])
            #elif value == "CONTACT_ID":
            #    values.append(call_list_method(token, "crm.contact.get", fields=[int(deal[value])])[1]['NAME'])
        deals.append(values)
    context = {"fields": fields, "deals": deals,}
    return render(request, 'GavlanDealApp/last_deals.html', context)

@main_auth(on_cookies=True)
def add_deal(request):
    token = request.bitrix_user_token
    if request.method == "POST":
        token.call_api_method('crm.deal.add', {"fields": request.POST.dict()})
    return render(request, 'GavlanDealApp/add_deal.html')

@main_auth(on_cookies=True)
def generate_qr(request):
    token = request.bitrix_user_token
    if request.method == "POST":
        id = int(request.POST.dict()['id'])
        product = token.call_api_method('crm.product.get', params={"id": id})['result']

        name = product['NAME']
        descritpion = product['DESCRIPTION']
        image_url = "https://b24-klxrii.bitrix24.ru" + str(product['PROPERTY_44'][0]['value']['downloadUrl'])
        response = requests.get(image_url, stream=True)
        image = Image.open(io.BytesIO(response.content))
        image_path = rf"GavlanDealApp\product_images\{id}.webp"
        image.save(rf"GavlanDealApp\static\GavlanDealApp\product_images\{id}.webp")

        product_db, _ = Product.objects.get_or_create(product_id=id, defaults = {'product_name': name, 'product_description': descritpion, 'product_image': image_path})
        product_db.save()

        s = get_random_string(16)
        signer = Signer(salt=s)
        signed_id = signer.sign_object(id)
        qr_link = f"localhost:8003/product?id={signed_id}&s={s}"

        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_link)
        qr.make(fit=True)
        qr_code_image = qr.make_image(fill_color="black", back_color="white")

        buffer = io.BytesIO()
        qr_code_image.save(buffer, format="PNG")
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        context = {"qr_link": qr_link, "qr_code": f"data:image/png;base64,{qr_base64}" }

        return render(request, 'GavlanDealApp/generate_qr.html', context)
    return render(request, 'GavlanDealApp/generate_qr.html')

def product(request):
    signer = Signer(salt=request.GET.get('s'))
    id = int(signer.unsign_object(request.GET.get('id')))
    product = Product.objects.get(product_id=id)
    objects = Product.objects.all()
    for object in objects:
        print(object.product_image, object.product_name, object.product_id)
    name = product.product_name
    description = product.product_description
    image = product.product_image
    context = {"name": name, "description": description, "image": image}
    return render(request, 'GavlanDealApp/product.html', context)


@main_auth(on_cookies=True)
def product_list(request):
    token = request.bitrix_user_token
    fields = ["ID", "Наименование"]
    product_types = ["ID", "TITLE"]
    full_products = sorted(token.call_list_method("crm.product.list"), key=itemgetter('ID'), reverse=True)
    products = []
    for product in full_products:
        values = []
        id = product["ID"]
        name = product["NAME"]
        values.append(id)
        values.append(name)
        products.append(values)
    context = {"fields": fields, "products": products,}
    return render(request, 'GavlanDealApp/product_list.html', context)

# def add_calls(employees, token):
#     for employee in employees:
#         calls = randint(1, 10)
#         call_time = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
#         for i in range (calls):
#             employee_id = employee['id']
#             call_start = token.call_api_method("telephony.externalcall.register",
#                                                params={"PHONE_NUMBER":"1", "USER_ID": employee_id, "CALL_START_DATE": call_time, "TYPE": 1})['result']
#             call_end = token.call_api_method("telephony.externalcall.finish", params={"CALL_ID": call_start["CALL_ID"], "USER_ID": employee_id, "DURATION": randint(1, 200)})
#     return call_end


#@lru_cache(maxsize=None)
def get_heads (department, heads, token, departments):
    department_id = department['ID']
    department_heads = token.call_list_method("im.department.managers.get", fields={'ID': [int(department_id)], "USER_DATA": "Y"})
    if department_heads:
        heads += department_heads[department_id]
    if 'PARENT' in department:
        return get_heads([d for d in departments if d['ID'] == department['PARENT']][0], heads, token, departments)
    else:
        return heads
@main_auth(on_cookies=True)
def employees(request):
    token = request.bitrix_user_token
    departments = token.call_api_method("department.get")['result']

    employees = []
    for department in departments:
        all_heads = get_heads(department, [], token, departments)
        employee_list = token.call_list_method("im.department.employees.get",  fields={"ID": [int(department['ID'])], "USER_DATA": "Y"})
        if employee_list:
            for employee in employee_list[department['ID']]:
                employee_id = employee['id']
                employee_name = employee['name']
                employee_heads = []
                for head in all_heads:
                    if head['id'] != employee_id:
                        employee_heads.append(head['name'])
                duplicates = list(filter(lambda e: e['id'] == employee_id, employees))
                if duplicates:
                    duplicates[0]['heads'].append(employee_heads)
                elif employee['active']:
                    now = datetime.now(timezone.utc)
                    earliest = (now - timedelta(hours=24)).isoformat()
                    calls = token.call_list_method("voximplant.statistic.get", fields={"FILTER": {">CALL_DURATION": 60, ">CALL_START_DATE": earliest}})
                    employees.append({"name": employee_name, "heads": employee_heads, "id": employee_id, "calls": len(calls),})

    # add_calls(employees, token)
    employee_list = []
    for employee in employees:
        employee_list.append([employee['name'], employee['heads'], employee['calls'], employee["id"]])

    context = {"employees": employee_list, }
    return render(request, 'GavlanDealApp/employees.html', context)

@main_auth(on_cookies=True)
def on_map (request):
    token = request.bitrix_user_token
    companies = []
    addresses = token.call_api_method("crm.address.list")['result']
    for address in addresses:
        if address['ENTITY_TYPE_ID'] == '8':
            company = token.call_api_method("crm.company.get", params={'id': address['ENTITY_ID']})['result']['TITLE']
            full_address = f"{address['COUNTRY']}, {address['CITY']}, {address['ADDRESS_1']}"
            companies.append([full_address, company])
    context = {'companies': companies}
    return render(request, 'GavlanDealApp/on_map.html', context)


def get_data(file):
    extension = splitext(file.name)[1].lower()
    if extension == ".csv":
        csv_file = io.TextIOWrapper(file.file, encoding='utf-8')
        result = csv.reader(csv_file)
        return result
    elif extension == ".xlsx":
        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), read_only=True)
        result = [list(row) for row in wb.active.iter_rows(values_only=True)]
        return result
@main_auth(on_cookies=True)
def upload_contacts(request):
    token = request.bitrix_user_token
    if request.method == "POST":
        file = request.FILES.get('file')
        contact_list = get_data(file)
        companies = token.call_list_method("crm.company.list", fields={"select":["ID", "TITLE"]})
        # names = ["NAME", "LAST_NAME", "PHONE", "EMAIL", "COMPANY_ID"]
        contacts = []
        for contact in contact_list:
            company_id = next((company for company in companies if company["TITLE"] == contact[4]), None)['ID']
            contact_dict = {"NAME": contact[0], "LAST_NAME": contact[1], "PHONE": [{"VALUE": contact[2]}], "EMAIL": [{"VALUE": contact[3]}], "COMPANY_ID": company_id}
            contacts.append(("crm.contact.add", {'fields': contact_dict}))
        _batch_api_call(contacts, token, token.call_api_method)
        return render(request, 'GavlanDealApp/upload_contacts.html', {"success": True})
    return render(request, 'GavlanDealApp/upload_contacts.html', )


def get_file(data, type):
    if type == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerows(data)
        file = buffer.getvalue()
        response = HttpResponse(file, content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="contacts.csv"'
        return response
    elif type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        buffer = io.BytesIO()
        wb.save(buffer)
        file = buffer.getvalue()
        response = HttpResponse(file, content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="contacts.xlsx"'
        return response
@main_auth(on_cookies=True)
def download_contacts(request):
    token = request.bitrix_user_token
    if request.method == "POST":
        file_type = request.POST.get("file_type")
        companies = token.call_list_method("crm.company.list", fields={"select": ["ID", "TITLE"]})
        contact_list = token.call_list_method("crm.contact.list", fields={"select": ["NAME", "LAST_NAME", "PHONE", "EMAIL", "COMPANY_ID"]})
        contacts = []
        for contact in contact_list:
            company_title = next((company for company in companies if company["ID"] == contact['COMPANY_ID']), None)['TITLE']
            contact_row = [contact["NAME"], contact["LAST_NAME"], contact["PHONE"][0]["VALUE"], contact["EMAIL"][0]["VALUE"], company_title]
            contacts.append(contact_row)
        file = get_file(contacts, file_type)
        return file
    return render(request, 'GavlanDealApp/download_contacts.html')

